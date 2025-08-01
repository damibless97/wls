from fastapi import FastAPI, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from openpyxl import Workbook, load_workbook
import os

app = FastAPI()

# Allow frontend from GitHub Pages
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5500"],  # Change to your actual GitHub Pages URL
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = "waitlist.xlsx"

@app.post("/join")
async def join_waitlist(email: str = Form(...)):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Email"])  # Header

    ws.append([email])
    wb.save(EXCEL_FILE)

    return JSONResponse(content={"message": "Success", "count": ws.max_row - 1})

@app.get("/count")
async def get_count():
    if not os.path.exists(EXCEL_FILE):
        return {"count": 0}
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return {"count": ws.max_row - 1}
